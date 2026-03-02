from aiogram import Router, F
from aiogram.types import Message, ReplyKeyboardMarkup, KeyboardButton
from datetime import datetime
import os

import database as db

router = Router()

def google_available():
return os.path.exists(“google_credentials.json”)

async def get_gsheets_client():
import gspread
from google.oauth2.service_account import Credentials
scopes = [
“https://www.googleapis.com/auth/spreadsheets”,
“https://www.googleapis.com/auth/drive”
]
creds = Credentials.from_service_account_file(“google_credentials.json”, scopes=scopes)
return gspread.authorize(creds)

# ─── МЕНЮ ЭКСПОРТА ───────────────────────────────────────────────────────────

@router.message(F.text == “📤 Экспорт в Excel”)
async def export_menu(message: Message):
if google_available():
buttons = [
[KeyboardButton(text=“📊 Экспорт в Google Таблицы”)],
[KeyboardButton(text=“📥 Скачать Excel файл”)],
[KeyboardButton(text=“🔙 Главное меню”)],
]
else:
buttons = [
[KeyboardButton(text=“📥 Скачать Excel файл”)],
[KeyboardButton(text=“⚙️ Как подключить Google Таблицы”)],
[KeyboardButton(text=“🔙 Главное меню”)],
]
await message.answer(“Выбери формат:”, reply_markup=ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True))

# ─── ИНСТРУКЦИЯ GOOGLE ───────────────────────────────────────────────────────

@router.message(F.text == “⚙️ Как подключить Google Таблицы”)
async def google_setup_help(message: Message):
text = (
“📋 <b>Подключение Google Таблиц — пошагово:</b>\n\n”
“<b>Шаг 1.</b> Зайди на console.cloud.google.com\n\n”
“<b>Шаг 2.</b> Создай проект (кнопка вверху слева)\n\n”
“<b>Шаг 3.</b> APIs & Services → Library → найди и включи:\n”
“   • Google Sheets API\n”
“   • Google Drive API\n\n”
“<b>Шаг 4.</b> APIs & Services → Credentials →\n”
“   Create Credentials → Service Account\n\n”
“<b>Шаг 5.</b> На странице Service Account:\n”
“   Keys → Add Key → Create new key → JSON\n\n”
“<b>Шаг 6.</b> Скачанный файл переименуй в:\n”
“   <code>google_credentials.json</code>\n”
“   и положи рядом с <code>bot.py</code>\n\n”
“<b>Шаг 7.</b> Установи библиотеки:\n”
“   <code>pip install gspread google-auth</code>\n\n”
“✅ После этого перезапусти бота — появится кнопка экспорта в Google Таблицы!”
)
from handlers.common import MAIN_MENU
await message.answer(text, parse_mode=“HTML”, reply_markup=MAIN_MENU)

# ─── ЭКСПОРТ В GOOGLE ТАБЛИЦЫ ────────────────────────────────────────────────

@router.message(F.text == “📊 Экспорт в Google Таблицы”)
async def export_google_sheets(message: Message):
if not google_available():
await message.answer(“❌ Файл google_credentials.json не найден.\nНажми ‘⚙️ Как подключить Google Таблицы’”)
return

```
await message.answer("⏳ Создаю Google Таблицу, подожди...")

try:
    now = datetime.now()
    date_from = now.replace(day=1).strftime("%Y-%m-%d")
    month_name = now.strftime("%B %Y")

    encashments = await db.get_encashments(date_from=date_from)
    expenses = await db.get_expenses(date_from=date_from)
    stats = await db.get_stats(date_from=date_from)
    points = await db.get_all_points()

    gc = await get_gsheets_client()

    # Открываем или создаём таблицу
    title = f"Духи — {month_name}"
    try:
        sh = gc.open(title)
    except Exception:
        sh = gc.create(title)
        sh.share(None, perm_type='anyone', role='reader')

    total_revenue = sum(s["total_revenue"] or 0 for s in stats)
    total_profit = sum(s["total_profit"] or 0 for s in stats)
    total_bottles = sum(s["total_bottles"] or 0 for s in stats)
    total_expenses = sum(e["amount"] for e in expenses)
    net_profit = total_profit - total_expenses

    BLUE = {"red": 0.17, "green": 0.36, "blue": 0.83}
    WHITE = {"red": 1, "green": 1, "blue": 1}

    # ── Лист: Сводка ────────────────────────────────────────────────────
    try:
        ws1 = sh.worksheet("Сводка")
        ws1.clear()
    except Exception:
        ws1 = sh.add_worksheet("Сводка", rows=50, cols=6)

    ws1.update("A1", [
        [f"📊 Отчёт за {month_name}", ""],
        ["", ""],
        ["Показатель", "Значение"],
        ["💵 Выручка (₽)", round(total_revenue)],
        ["💰 Прибыль до расходов (₽)", round(total_profit)],
        ["💸 Расходы (₽)", round(total_expenses)],
        ["✅ Чистая прибыль (₽)", round(net_profit)],
        ["📦 Продано флаконов", total_bottles],
        ["", ""],
        ["📍 По точкам", ""],
        ["Точка", "Выручка (₽)", "Прибыль (₽)", "Продано (шт.)"],
    ] + [
        [s["name"], round(s["total_revenue"] or 0), round(s["total_profit"] or 0), s["total_bottles"] or 0]
        for s in stats
    ])
    ws1.format("A1:B1", {"textFormat": {"bold": True, "fontSize": 14}})
    ws1.format("A3:B3", {"backgroundColor": BLUE, "textFormat": {"bold": True, "foregroundColor": WHITE}})
    ws1.format("A7:B7", {"textFormat": {"bold": True, "fontSize": 12}})
    ws1.format("A11:D11", {"backgroundColor": BLUE, "textFormat": {"bold": True, "foregroundColor": WHITE}})

    # ── Лист: Инкассации ────────────────────────────────────────────────
    try:
        ws2 = sh.worksheet("Инкассации")
        ws2.clear()
    except Exception:
        ws2 = sh.add_worksheet("Инкассации", rows=500, cols=8)

    enc_data = [["Дата", "Точка", "Сумма (₽)", "Продано (шт.)", "Прибыль (₽)", "Примечание"]]
    for e in encashments:
        enc_data.append([e["date"], e["point_name"], round(e["amount"]),
                         e["bottles_sold"], round(e["my_profit"]), e["note"] or ""])
    ws2.update("A1", enc_data)
    ws2.format("A1:F1", {"backgroundColor": BLUE, "textFormat": {"bold": True, "foregroundColor": WHITE}})

    # ── Лист: Расходы ───────────────────────────────────────────────────
    try:
        ws3 = sh.worksheet("Расходы")
        ws3.clear()
    except Exception:
        ws3 = sh.add_worksheet("Расходы", rows=500, cols=5)

    exp_data = [["Дата", "Категория", "Сумма (₽)", "Примечание"]]
    for exp in expenses:
        exp_data.append([exp["date"], exp["category"], round(exp["amount"]), exp["note"] or ""])
    ws3.update("A1", exp_data)
    ws3.format("A1:D1", {"backgroundColor": BLUE, "textFormat": {"bold": True, "foregroundColor": WHITE}})

    # ── Лист: Остатки ───────────────────────────────────────────────────
    try:
        ws4 = sh.worksheet("Остатки")
        ws4.clear()
    except Exception:
        ws4 = sh.add_worksheet("Остатки", rows=200, cols=5)

    stock_data = [["Точка", "Аромат", "Остаток (шт.)"]]
    for point in points:
        stock = await db.get_point_stock(point["id"])
        for item in stock:
            stock_data.append([point["name"], item["aroma_name"], item["quantity"]])
    ws4.update("A1", stock_data)
    ws4.format("A1:C1", {"backgroundColor": BLUE, "textFormat": {"bold": True, "foregroundColor": WHITE}})

    # Удаляем пустой Sheet1
    try:
        sh.del_worksheet(sh.worksheet("Sheet1"))
    except Exception:
        pass

    from handlers.common import MAIN_MENU
    await message.answer(
        f"✅ <b>Google Таблица готова!</b>\n\n"
        f"📊 Отчёт за {month_name}\n"
        f"💰 Чистая прибыль: <b>{net_profit:.0f}₽</b>\n\n"
        f"🔗 <a href='{sh.url}'>Открыть таблицу</a>",
        parse_mode="HTML",
        reply_markup=MAIN_MENU,
        disable_web_page_preview=True
    )

except Exception as e:
    from handlers.common import MAIN_MENU
    await message.answer(
        f"❌ Ошибка:\n<code>{str(e)}</code>",
        parse_mode="HTML",
        reply_markup=MAIN_MENU
    )
```

# ─── ЭКСПОРТ В EXCEL ─────────────────────────────────────────────────────────

@router.message(F.text == “📥 Скачать Excel файл”)
async def export_excel(message: Message):
await message.answer(“⏳ Формирую Excel файл…”)

```
import openpyxl
from openpyxl.styles import Font, PatternFill
from aiogram.types import FSInputFile

now = datetime.now()
date_from = now.replace(day=1).strftime("%Y-%m-%d")
month_name = now.strftime("%B_%Y")
filename = f"/tmp/report_{month_name}.xlsx"

encashments = await db.get_encashments(date_from=date_from)
expenses = await db.get_expenses(date_from=date_from)
stats = await db.get_stats(date_from=date_from)

wb = openpyxl.Workbook()
FILL = PatternFill("solid", fgColor="2B5DD4")
BOLD_WHITE = Font(color="FFFFFF", bold=True)

total_revenue = sum(s["total_revenue"] or 0 for s in stats)
total_profit = sum(s["total_profit"] or 0 for s in stats)
total_bottles = sum(s["total_bottles"] or 0 for s in stats)
total_expenses = sum(e["amount"] for e in expenses)
net_profit = total_profit - total_expenses

# Сводка
ws1 = wb.active
ws1.title = "Сводка"
ws1["A1"] = f"Отчёт за {now.strftime('%B %Y')}"
ws1["A1"].font = Font(bold=True, size=14)
for i, (label, value) in enumerate([
    ("Выручка (₽)", round(total_revenue)),
    ("Прибыль до расходов (₽)", round(total_profit)),
    ("Расходы (₽)", round(total_expenses)),
    ("Чистая прибыль (₽)", round(net_profit)),
    ("Продано флаконов", total_bottles),
], start=3):
    ws1[f"A{i}"] = label
    ws1[f"B{i}"] = value
    ws1[f"A{i}"].font = Font(bold=True)
ws1["B6"].font = Font(bold=True, size=12)

ws1["A9"] = "По точкам:"
ws1["A9"].font = Font(bold=True)
for col, h in enumerate(["Точка", "Выручка (₽)", "Прибыль (₽)", "Продано (шт.)"], 1):
    c = ws1.cell(row=10, column=col, value=h)
    c.fill = FILL
    c.font = BOLD_WHITE
for row, s in enumerate(stats, start=11):
    ws1.cell(row=row, column=1, value=s["name"])
    ws1.cell(row=row, column=2, value=round(s["total_revenue"] or 0))
    ws1.cell(row=row, column=3, value=round(s["total_profit"] or 0))
    ws1.cell(row=row, column=4, value=s["total_bottles"] or 0)
for col in ["A", "B", "C", "D"]:
    ws1.column_dimensions[col].width = 22

# Инкассации
ws2 = wb.create_sheet("Инкассации")
for col, h in enumerate(["Дата", "Точка", "Сумма (₽)", "Продано (шт.)", "Прибыль (₽)", "Примечание"], 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.fill = FILL
    c.font = BOLD_WHITE
for row, e in enumerate(encashments, start=2):
    ws2.cell(row=row, column=1, value=e["date"])
    ws2.cell(row=row, column=2, value=e["point_name"])
    ws2.cell(row=row, column=3, value=round(e["amount"]))
    ws2.cell(row=row, column=4, value=e["bottles_sold"])
    ws2.cell(row=row, column=5, value=round(e["my_profit"]))
    ws2.cell(row=row, column=6, value=e["note"] or "")
for col in ["A", "B", "C", "D", "E", "F"]:
    ws2.column_dimensions[col].width = 18

# Расходы
ws3 = wb.create_sheet("Расходы")
for col, h in enumerate(["Дата", "Категория", "Сумма (₽)", "Примечание"], 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.fill = FILL
    c.font = BOLD_WHITE
for row, exp in enumerate(expenses, start=2):
    ws3.cell(row=row, column=1, value=exp["date"])
    ws3.cell(row=row, column=2, value=exp["category"])
    ws3.cell(row=row, column=3, value=round(exp["amount"]))
    ws3.cell(row=row, column=4, value=exp["note"] or "")
for col in ["A", "B", "C", "D"]:
    ws3.column_dimensions[col].width = 20

wb.save(filename)
doc = FSInputFile(filename, filename=f"Отчёт_{month_name}.xlsx")
from handlers.common import MAIN_MENU
await message.answer_document(
    doc,
    caption=f"📊 Отчёт за {now.strftime('%B %Y')}\n💰 Чистая прибыль: {net_profit:.0f}₽",
    reply_markup=MAIN_MENU
)
os.remove(filename)
```
